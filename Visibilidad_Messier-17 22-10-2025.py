import pandas as pd
import numpy as np
from astropy.time import Time
from astropy.coordinates import SkyCoord, EarthLocation, AltAz, get_body, get_sun
from astropy import units as u
from astropy.utils import iers
from datetime import datetime, timedelta
import pytz
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from tkinter import Tk, filedialog

# =========================================
# Configurar IERS
# =========================================
iers.conf.auto_download = True
iers.conf.iers_auto_url = 'https://datacenter.iers.org/data/9/finals2000A.all'

# =========================================
# Mensaje de bienvenida
# =========================================
mensaje_intro = """
=========================================================== 
      BIENVENIDO AL PLANIFICADOR DE MARATONES MESSIER
===========================================================
Este programa genera una tabla optimizada de los objetos Messier
visibles desde tu ubicación y en el intervalo horario indicado.

Su objetivo es ayudarte a planificar la observación de forma 
eficiente, ordenando los objetos según su prioridad de observación,
para aprovechar al máximo la noche y evitar pérdidas por ocultación.

La tabla incluye además la distancia angular a la Luna, dato esencial 
para identificar posibles interferencias por brillo lunar y decidir 
qué objetos conviene observar cuando el cielo esté más oscuro.

===========================================================
"""
print(mensaje_intro)

# =========================================
# Seleccionar ruta donde guardar Excel
# =========================================
root = Tk()
root.withdraw()
root.attributes('-topmost', True)
print("Selecciona la carpeta donde deseas guardar la tabla de Excel...")
ruta_guardado = filedialog.askdirectory()
root.destroy()
if not ruta_guardado:
    print("⚠️ No seleccionaste ninguna carpeta. Se usará el escritorio por defecto.")
    ruta_guardado = os.path.join(os.path.expanduser("~"), "Desktop")

SALIDA_EXCEL = os.path.join(
    ruta_guardado,
    f"Visibilidad_Messier_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

# =========================================
# Ruta del archivo de entrada (Libro1.xlsx)
# =========================================
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Libro1.xlsx")

# =========================================
# Parámetros fijos
# =========================================
ALT_MIN = -0.566
ALERTA_UMBRAL_GRADOS = 15

# =========================================
# Cargar datos
# =========================================
df = pd.read_excel(EXCEL_PATH)
df = df.dropna(subset=[df.columns[1], df.columns[2], df.columns[4], df.columns[5], df.columns[6]])

ra_h = pd.to_numeric(df.iloc[:, 1], errors='coerce').fillna(0)
ra_m = pd.to_numeric(df.iloc[:, 2], errors='coerce').fillna(0)
dec_deg = pd.to_numeric(df.iloc[:, 4], errors='coerce').fillna(0)
dec_min = pd.to_numeric(df.iloc[:, 5], errors='coerce').fillna(0)
magnitudes = pd.to_numeric(df.iloc[:, 6], errors='coerce').fillna(np.nan)
constelaciones = df["Constelación"].fillna("—")

ra_deg = (ra_h + ra_m / 60.0) * 15.0
dec_deg_total = np.where(dec_deg >= 0, dec_deg + dec_min / 60.0, dec_deg - dec_min / 60.0)
coords = SkyCoord(ra=ra_deg.values * u.deg, dec=dec_deg_total * u.deg, frame='icrs')

# =========================================
# Entradas del usuario
# =========================================
latitud = float(input("Latitud (grados, +N, -S): "))
longitud = float(input("Longitud (grados, +E, -O): "))
altura_msnm = float(input("Altura sobre el nivel del mar (m): "))
fecha_str = input("Fecha observación (AAAA-MM-DD): ")

while True:
    try:
        tz_offset = float(input("Introduce tu zona horaria en número (ej. -5): "))
        LOCAL_TZ = pytz.FixedOffset(int(tz_offset * 60))
        break
    except ValueError:
        print("⚠️ Entrada inválida. Debe ser un número, ej. -5, +2, etc.")

hora_inicio_str = input("Hora de inicio de observación (HH:MM, 24h): ")
hora_fin_str = input("Hora de fin de observación (HH:MM, 24h): ")

h_i, m_i = map(int, hora_inicio_str.split(":"))
h_f, m_f = map(int, hora_fin_str.split(":"))

# =========================================
# Configuración de tiempos
# =========================================
local_tz = LOCAL_TZ
fecha_local = local_tz.localize(datetime.strptime(fecha_str, "%Y-%m-%d"))

inicio = fecha_local.replace(hour=h_i, minute=m_i, second=0)
fin = fecha_local.replace(hour=h_f, minute=m_f, second=0)
if fin <= inicio:
    fin += timedelta(days=1)

tiempos_locales_vent = pd.date_range(start=inicio, end=fin, freq='5min', tz=local_tz)
tiempos_utc_vent = [Time(t.astimezone(pytz.utc)) for t in tiempos_locales_vent]

start_all = inicio - timedelta(hours=12)
end_all = inicio + timedelta(hours=36)
grid_local_all = pd.date_range(start=start_all, end=end_all, freq='1min', tz=local_tz)
grid_utc_all = [Time(t.astimezone(pytz.utc)) for t in grid_local_all]

location = EarthLocation(lat=latitud * u.deg, lon=longitud * u.deg, height=altura_msnm * u.m)

# =========================================
# Calcular crepúsculos astronómicos
# =========================================
def calcular_crepusculo(fecha_base, location, local_tz, tipo="vespertino"):
    """Devuelve la hora local del crepúsculo astronómico vespertino o matutino"""
    pasos = pd.date_range(
        start=fecha_base.replace(hour=15, minute=0) if tipo == "vespertino" else fecha_base.replace(hour=0, minute=0),
        end=fecha_base + timedelta(days=1),
        freq="2min",
        tz=local_tz
    )
    sol_alturas = []
    for t in pasos:
        sol = get_sun(Time(t.astimezone(pytz.utc)))
        alt = sol.transform_to(AltAz(obstime=Time(t.astimezone(pytz.utc)), location=location)).alt.deg
        sol_alturas.append(alt)
    sol_alturas = np.array(sol_alturas)
    if tipo == "vespertino":
        idx = np.where(sol_alturas < -18)[0]
        if len(idx) > 0:
            return pasos[idx[0]].strftime("%H:%M")
    else:
        idx = np.where(sol_alturas > -18)[0]
        if len(idx) > 0:
            return pasos[idx[0]].strftime("%H:%M")
    return "—"

crep_vespertino = calcular_crepusculo(fecha_local, location, local_tz, "vespertino")
crep_matutino = calcular_crepusculo(fecha_local + timedelta(days=1), location, local_tz, "matutino")

# =========================================
# Posición promedio de la Luna
# =========================================
mid_time = Time(inicio.astimezone(pytz.utc)) + (Time(fin.astimezone(pytz.utc)) - Time(inicio.astimezone(pytz.utc))) / 2
moon = get_body("moon", mid_time, location=location)
moon_altaz = moon.transform_to(AltAz(obstime=mid_time, location=location, pressure=0*u.hPa))

# =========================================
# Cálculos principales
# =========================================
resultados = []
for i, c in enumerate(coords):
    altaz_vent = c.transform_to(AltAz(obstime=tiempos_utc_vent, location=location))
    alt_vent = altaz_vent.alt.deg
    vis_indices_vent = np.where(alt_vent > ALT_MIN)[0]
    tiempo_visible_segundos = 0
    if len(vis_indices_vent) > 0:
        t_vent_salida = tiempos_locales_vent[vis_indices_vent[0]]
        t_vent_puesta = tiempos_locales_vent[vis_indices_vent[-1]]
        tiempo_visible_segundos = max(0, (t_vent_puesta - t_vent_salida).total_seconds())

    altaz_all = c.transform_to(AltAz(obstime=grid_utc_all, location=location))
    alt_all = altaz_all.alt.deg
    vis_bool = alt_all > ALT_MIN
    rises = np.where((vis_bool[1:] & ~vis_bool[:-1]))[0] + 1
    sets = np.where((~vis_bool[1:] & vis_bool[:-1]))[0] + 1
    t_salida_real = None
    t_puesta_real = None
    t_culminacion = None
    if len(rises) > 0 and len(sets) > 0:
        sets_arr = np.array(sets)
        for r in rises:
            s_candidates = sets_arr[sets_arr > r]
            if s_candidates.size > 0:
                s = s_candidates[0]
                t_salida_real = grid_local_all[r].to_pydatetime()
                t_puesta_real = grid_local_all[s].to_pydatetime()
                alt_window = alt_all[r:s+1]
                max_idx = np.argmax(alt_window)
                t_culminacion = grid_local_all[r + max_idx].to_pydatetime()
                break

    minutos_visible = int(tiempo_visible_segundos // 60)
    tiempo_h_m = f"{minutos_visible // 60}:{minutos_visible % 60:02d}"

    ra_hms = c.ra.hms
    dec_dms = c.dec.signed_dms
    ra_str = f"{int(ra_hms.h):02d}:{int(ra_hms.m):02d}:{ra_hms.s:.1f}"
    dec_sign = "+" if dec_dms.sign >= 0 else "-"
    dec_str = f"{dec_sign}{abs(int(dec_dms.d)):02d}:{int(abs(dec_dms.m)):02d}"

    puesta_ts = tiempos_locales_vent[vis_indices_vent[-1]].timestamp() if len(vis_indices_vent) > 0 else np.inf

    obj_altaz = c.transform_to(AltAz(obstime=mid_time, location=location, pressure=0*u.hPa))
    distancia_luna = obj_altaz.separation(moon_altaz).deg

    resultados.append({
        "Objeto": df.iloc[i, 0],
        "Constelación": constelaciones.iloc[i],
        "RA (HH:MM:SS)": ra_str,
        "Magnitud total": magnitudes.iloc[i],
        "DEC (±DD:MM)": dec_str,
        "Hora_salida_real": t_salida_real,
        "Hora_puesta_real": t_puesta_real,
        "Hora_culminación_real": t_culminacion,
        "Tiempo_visible_hm": tiempo_h_m,
        "Minutos_visible": minutos_visible,
        "Puesta_window_ts": puesta_ts,
        "Distancia_Luna (°)": round(distancia_luna, 2)
    })

# =========================================
# Construir tabla final (igual que antes)
# =========================================
res_df = pd.DataFrame(resultados)
visibles = res_df[res_df["Minutos_visible"] > 0].copy()
no_visibles = res_df[res_df["Minutos_visible"] == 0].copy()

visibles["Tiempo_restante_min"] = visibles.apply(
    lambda row: (row["Hora_puesta_real"] - inicio).total_seconds()/60 if row["Hora_puesta_real"] else np.inf, axis=1
)

visibles['idx_df'] = visibles.index
visibles.sort_values(by="Tiempo_restante_min", inplace=True)
visibles.reset_index(drop=True, inplace=True)

ordered_indices = []
remaining = visibles.index.tolist()
while remaining:
    remaining.sort(key=lambda r: visibles.loc[r, "Tiempo_restante_min"])
    idx = remaining[0]
    block = [idx]
    remaining.remove(idx)
    while len(block) < 5 and remaining:
        last_idx_df = visibles.loc[block[-1], 'idx_df']
        last_coord = coords[last_idx_df]
        closest_idx = min(
            remaining,
            key=lambda r: coords[visibles.loc[r, 'idx_df']].separation(last_coord).deg
        )
        block.append(closest_idx)
        remaining.remove(closest_idx)
    ordered_indices.extend(block)

visibles = visibles.loc[ordered_indices].reset_index(drop=True)
visibles["Orden sugerido"] = visibles.index + 1
visibles["Alerta_horizonte"] = visibles.apply(
    lambda row: "Sí" if row["Tiempo_restante_min"] <= ALERTA_UMBRAL_GRADOS*6 else "No", axis=1
)

def fmt_time(dt):
    return dt.strftime("%H:%M") if (dt is not None and not pd.isna(dt)) else "—"

visibles["Hora de salida"] = visibles["Hora_salida_real"].apply(fmt_time)
visibles["Hora de puesta"] = visibles["Hora_puesta_real"].apply(fmt_time)
visibles["Hora de culminación"] = visibles["Hora_culminación_real"].apply(fmt_time)
visibles["Tiempo visible (h:m)"] = visibles["Tiempo_visible_hm"]

tabla_final = visibles[[
    "Orden sugerido", "Objeto", "Constelación", "Magnitud total", "RA (HH:MM:SS)", "DEC (±DD:MM)",
    "Hora de salida", "Hora de puesta", "Hora de culminación", "Tiempo visible (h:m)",
    "Distancia_Luna (°)", "Alerta_horizonte"
]]

mensaje_tiempo_visible = f"Tiempo visible entre {hora_inicio_str} y {hora_fin_str}"
explicacion = pd.DataFrame([{
    "Orden sugerido": "—",
    "Objeto": "",
    "Constelación": "—",
    "Magnitud total": "",
    "RA (HH:MM:SS)": "—",
    "DEC (±DD:MM)": "—",
    "Hora de salida": "—",
    "Hora de puesta": "—",
    "Hora de culminación": "—",
    "Tiempo visible (h:m)": mensaje_tiempo_visible,
    "Distancia_Luna (°)": "Distancia angular a la Luna (grados)",
    "Alerta_horizonte": "‘Sí’ indica objeto muy cercano al horizonte oeste (<15°)"
}])

if not no_visibles.empty:
    no_visibles["Tiempo visible (h:m)"] = "0:00"
    no_visibles["Hora de salida"] = "—"
    no_visibles["Hora de puesta"] = "—"
    no_visibles["Hora de culminación"] = "—"
    no_visibles["Orden sugerido"] = "—"
    no_visibles["Alerta_horizonte"] = "—"
    no_visibles["Distancia_Luna (°)"] = "—"
    no_visibles["Objeto"] = no_visibles["Objeto"] + " (NO VISIBLE)"
    no_visibles["Constelación"] = "—"

final_tabla = pd.concat([explicacion, tabla_final, no_visibles[tabla_final.columns]], ignore_index=True)

# =========================================
# Guardar Excel y añadir datos del usuario + crepúsculos
# =========================================
final_tabla.to_excel(SALIDA_EXCEL, index=False)
wb = load_workbook(SALIDA_EXCEL)
ws = wb.active

# Ajuste automático de columnas
for col in ws.columns:
    max_len_data = 0
    col_letter = get_column_letter(col[0].column)
    header = ws[col_letter + "1"].value
    header_len = len(str(header)) if header else 0
    for cell in col:
        if cell.value:
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            max_len_data = max(max_len_data, len(str(cell.value)))
    ancho_final = max(header_len * 1.1, min(max_len_data * 1.1, 25))
    ws.column_dimensions[col_letter].width = ancho_final

# Añadir los datos del usuario en la parte superior (A-C combinadas)
info_usuario = [
    f"Latitud: {latitud}°   Longitud: {longitud}°   Altura: {altura_msnm} m",
    f"Fecha de observación: {fecha_str}",
    f"Zona horaria: {tz_offset:+}   Hora inicio: {hora_inicio_str}   Hora fin: {hora_fin_str}",
    f"Crepúsculo astronómico vespertino: {crep_vespertino} (aprox)",
    f"Crepúsculo astronómico matutino: {crep_matutino} (aprox)"
]

ws.insert_rows(1, amount=7)
for i, texto in enumerate(info_usuario, start=1):
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
    celda = ws.cell(row=i, column=1)
    celda.value = texto
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.font = Font(bold=True)

wb.save(SALIDA_EXCEL)
print(f"\n✅ Tabla generada y guardada en: {SALIDA_EXCEL}")
